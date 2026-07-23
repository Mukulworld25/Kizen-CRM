import glob
import os
import openpyxl
import json

downloads = 'C:/Users/admin/Downloads'
files = [
    'Leads for Kizen.xlsx',
    'Data for Preeti.xlsx',
    'My students.xlsx',
    'Leads by Lakshaya Ma\'am.xlsx',
    'Fees Tracker.xlsx',
    '_Fee Structure- Kizen.xlsx'
]

summary = {}

for fname in files:
    fpath = os.path.join(downloads, fname)
    if not os.path.exists(fpath):
        print(f"MISSING: {fname}")
        continue
    
    wb = openpyxl.load_workbook(fpath, data_only=True)
    summary[fname] = {}
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = []
        rows = list(ws.iter_rows(values_only=True))
        
        # Find first non-empty row as header
        header_idx = -1
        for idx, r in enumerate(rows[:5]):
            if any(r):
                headers = [str(cell).strip() if cell is not None else '' for cell in r]
                header_idx = idx
                break
        
        non_empty_rows = sum(1 for r in rows if any(r))
        summary[fname][sname] = {
            'total_rows': len(rows),
            'data_rows': max(0, non_empty_rows - 1),
            'header_line': header_idx,
            'headers': [h for h in headers if h][:10] # first 10 headers
        }

print("\n=== WORKBOOK HEADERS & STRUCTURE AUDIT ===")
for fname, sheets in summary.items():
    print(f"\n📂 File: {fname}")
    for sname, info in sheets.items():
        print(f"  📄 Tab: [{sname.strip()}] -> {info['data_rows']} data rows | Headers: {info['headers']}")

with open('workbook_audit.json', 'w') as out:
    json.dump(summary, out, indent=2)
