import os
import glob
import re
import openpyxl
from supabase import create_client, Client

SU = 'https://zmqvjtenuxlvwfopfroc.supabase.co'
SK = 'sb_publishable_ezZ7_UyticpI7aeLKY3Rew_i_FaZBDh'
supabase: Client = create_client(SU, SK)

def clean_phone(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) >= 10:
        # Take the last 10 digits for Indian standard numbers
        return digits[-10:]
    return None

def clean_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    if s.lower() in ('none', 'null', 'nan', 'undefined', 'n/a'):
        return ''
    return s

def map_source(raw):
    s = clean_str(raw).lower()
    if 'insta' in s: return 'instagram'
    if 'face' in s or 'fb' in s: return 'facebook'
    if 'walk' in s: return 'walk_in'
    if any(k in s for k in ['ref', 'simrat', 'preeti', 'lakshay', 'aadya', 'friend']): return 'referral'
    if 'web' in s or 'site' in s: return 'website'
    if 'wa' in s or 'whatsapp' in s or 'sensy' in s: return 'whatsapp'
    if any(k in s for k in ['college', 'school', 'pu', 'cuet', '12th']): return 'college_visit'
    return 'other'

def map_status(raw):
    s = clean_str(raw).lower()
    if any(k in s for k in ['enrol', 'admit', 'paid', 'joined', 'converted']): return 'enrolled'
    if any(k in s for k in ['close', 'lost', 'not int', 'unpicked', 'reject', 'wrong']): return 'lost'
    if any(k in s for k in ['contact', 'visit', 'demo', 'intrest', 'follow', 'call', 'talk']): return 'contacted'
    return 'new'

def map_temperature(raw, tab_name=""):
    s = (clean_str(raw) + " " + tab_name).lower()
    if 'hot' in s: return 'hot'
    if 'warm' in s: return 'warm'
    if 'cold' in s: return 'cold'
    return 'warm'

def main():
    print("=== MASTER ZERO-ERROR MULTI-WORKBOOK INGESTION ENGINE ===")

    # Authenticate as Shivam Owner
    auth_resp = supabase.auth.sign_in_with_password({
        "email": "shivam.kizen.test@gmail.com",
        "password": "Shivam@123"
    })
    print(f"Authenticated as: {auth_resp.user.email}")

    # Fetch counselors
    users_resp = supabase.from_('users').select('id, name, email, role').execute()
    users_data = users_resp.data or []
    
    aadya_id = next((u['id'] for u in users_data if u['email'] == 'counselor1@kizen.edu'), None)
    lakshaya_id = next((u['id'] for u in users_data if u['email'] == 'lakshaya@kizen.edu'), None)
    preeti_id = next((u['id'] for u in users_data if u['email'] == 'reception@kizen.edu'), None)

    print(f"Counselor IDs: Aadya={aadya_id}, Lakshaya={lakshaya_id}, Preeti={preeti_id}")

    downloads = 'C:/Users/admin/Downloads'
    files = [
        'Leads for Kizen.xlsx',
        'Data for Preeti.xlsx',
        'My students.xlsx',
        'Leads by Lakshaya Ma\'am.xlsx',
        'Fees Tracker.xlsx',
        '_Fee Structure- Kizen.xlsx'
    ]

    leads_by_mobile = {}
    leads_by_name = {}

    total_raw_rows = 0

    for fname in files:
        fpath = os.path.join(downloads, fname)
        if not os.path.exists(fpath):
            print(f"SKIP (File not found): {fname}")
            continue

        print(f"\nProcessing File: {fname}")
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)

        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Identify headers
            header_idx = -1
            name_col = -1
            phone_col = -1
            city_col = -1
            course_col = -1
            status_col = -1
            notes_col = -1
            date_col = -1
            source_col = -1

            for idx, r in enumerate(rows[:5]):
                row_str = [str(c).lower() if c is not None else '' for c in r]
                if any('name' in c or 'contact' in c or 'mobile' in c or 'phone' in c for c in row_str):
                    header_idx = idx
                    for col_i, h in enumerate(row_str):
                        if 'name' in h and name_col == -1: name_col = col_i
                        elif any(k in h for k in ['contact', 'mobile', 'phone', 'number', 'whatsapp']) and phone_col == -1: phone_col = col_i
                        elif any(k in h for k in ['city', 'location', 'address', 'state']) and city_col == -1: city_col = col_i
                        elif any(k in h for k in ['class', 'course', 'qualificat', 'subject', 'program']) and course_col == -1: course_col = col_i
                        elif any(k in h for k in ['status', 'stage']) and status_col == -1: status_col = col_i
                        elif any(k in h for k in ['remark', 'note', 'comment', 'detail']) and notes_col == -1: notes_col = col_i
                        elif any(k in h for k in ['date', 'time', 'followup']) and date_col == -1: date_col = col_i
                        elif 'source' in h and source_col == -1: source_col = col_i
                    break

            if name_col == -1 and phone_col == -1:
                # Default positional fallback
                name_col = 1 if len(rows[0]) > 1 else 0
                phone_col = 2 if len(rows[0]) > 2 else -1

            sheet_count = 0
            for r in rows[header_idx + 1:]:
                if not any(r): continue
                total_raw_rows += 1

                raw_name = clean_str(r[name_col]) if 0 <= name_col < len(r) else ''
                raw_phone = r[phone_col] if 0 <= phone_col < len(r) else None
                phone = clean_phone(raw_phone)

                if not raw_name and not phone:
                    continue

                if raw_name.lower() in ('name', 'names', 'contact no', 'lead date', 'total', 's.no', 'sr.no'):
                    continue

                city = clean_str(r[city_col]) if 0 <= city_col < len(r) else ''
                course = clean_str(r[course_col]) if 0 <= course_col < len(r) else ''
                status_raw = clean_str(r[status_col]) if 0 <= status_col < len(r) else ''
                notes_raw = clean_str(r[notes_col]) if 0 <= notes_col < len(r) else ''
                date_raw = clean_str(r[date_col]) if 0 <= date_col < len(r) else ''
                source_raw = clean_str(r[source_col]) if 0 <= source_col < len(r) else ''

                # Counselor assignment logic
                assigned_counselor = preeti_id if 'preeti' in fname.lower() else (
                    lakshaya_id if 'lakshaya' in fname.lower() or 'lakshaya' in sname.lower() else aadya_id
                )

                status = map_status(status_raw)
                source = map_source(source_raw or sname)
                temp = map_temperature(notes_raw, sname)

                combined_notes = f"[{fname} -> {sname}] {notes_raw} {course}".strip()

                lead_data = {
                    'full_name': raw_name or f"Lead {phone}",
                    'mobile': phone or '9999999999',
                    'city': city or None,
                    'status': status,
                    'source': source,
                    'temperature': temp,
                    'notes': combined_notes,
                    'assigned_counselor_id': assigned_counselor
                }

                if phone and phone != '9999999999':
                    if phone in leads_by_mobile:
                        # Merge with existing lead
                        existing = leads_by_mobile[phone]
                        existing['notes'] += f" | {combined_notes}"
                        if status == 'enrolled': existing['status'] = 'enrolled'
                        if temp == 'hot': existing['temperature'] = 'hot'
                    else:
                        leads_by_mobile[phone] = lead_data
                else:
                    name_key = raw_name.lower()
                    if name_key in leads_by_name:
                        existing = leads_by_name[name_key]
                        existing['notes'] += f" | {combined_notes}"
                    else:
                        leads_by_name[name_key] = lead_data

                sheet_count += 1

            print(f"  Tab [{sname.strip()}]: Processed {sheet_count} valid lead records")

    master_leads = list(leads_by_mobile.values()) + list(leads_by_name.values())

    print(f"\n================ SUMMARY ================")
    print(f"Total Raw Spreadsheet Rows Read: {total_raw_rows}")
    print(f"Total Unique Master Leads Deduplicated: {len(master_leads)}")

    # Batch upsert to Supabase
    print("\nBatch upserting master leads to Supabase DB...")
    batch_size = 100
    inserted_count = 0

    for i in range(0, len(master_leads), batch_size):
        batch = master_leads[i:i+batch_size]
        res = supabase.from_('leads').upsert(batch, on_conflict='mobile').execute()
        inserted_count += len(batch)
        print(f"  Upserted batch {i//batch_size + 1} ({inserted_count}/{len(master_leads)})")

    print(f"\n✅ SUCCESS! All {len(master_leads)} Master Leads Ingested into Supabase Database!")

if __name__ == '__main__':
    main()
